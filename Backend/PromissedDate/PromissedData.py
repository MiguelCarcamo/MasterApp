import pandas as pd
from datetime import datetime
import asyncio

class MainPromissedDate:
    __Book = None
    __SheetBoot = None
    __pathOrigen = None
    __row = 0
    __rowDatos = 0
    __col = 0
    __Promissed = None
    __Assigned = None
    __AssignedWeeks = None


    # constructor
    def __init__(self, path1, SheetName, Plant, Promissed, Assigned, AssignedWeeks):
        import openpyxl
        self.__pathOrigen = r"%s"%path1
        self.__Book = openpyxl.load_workbook(self.__pathOrigen, data_only=True)
        self.__SheetBoot = self.__Book[SheetName]
        self.__row = self.__SheetBoot.max_row
        self.__col = self.__SheetBoot.max_column
        self.__Promissed = Promissed
        self.__Assigned = Assigned
        self.__AssignedWeeks = AssignedWeeks
    
    def PositionTitle(self):
        List = {}
        for x in range(self.__row):
            r = x + 1
            if self.__SheetBoot.cell(row=r, column= 1).value == "Title":
                self.__rowDatos = r + 1 
                for z in range(self.__col):
                    c = z + 1
                    if self.__SheetBoot.cell(row= r, column= c).value == "Po-Item":
                        List['Po-Item'] = str(c)
                    if self.__SheetBoot.cell(row= r, column= c).value == "Cut #":
                        List['Cut #'] = str(c)
                    if self.__SheetBoot.cell(row= r, column= c).value == self.__AssignedWeeks:
                        List[self.__AssignedWeeks] = str(c)
                    if self.__SheetBoot.cell(row= r, column= c).value == self.__Assigned:
                        List[self.__Assigned] = str(c)
                    if self.__SheetBoot.cell(row= r, column= c).value == self.__Promissed:
                        List[self.__Promissed] = str(c)
        return List
    
    def AddInfo(self):
        Data = {}
        Data['Info'] = []
        VarPositionTitle = self.PositionTitle()
        for r in range(self.__rowDatos, self.__row + 1):
            if self.__SheetBoot.cell(row= r, column= int(VarPositionTitle['Po-Item'])).value:
                Po = str(self.__SheetBoot.cell(row= r, column= int(VarPositionTitle['Po-Item'])).value)
                Cut = str(self.__SheetBoot.cell(row= r, column= int(VarPositionTitle['Cut #'])).value)
                try:
                    NumAssigned = str(self.__SheetBoot.cell(row= r, column= int(VarPositionTitle[self.__AssignedWeeks])).value)
                except:
                    NumAssigned = ""
                try:
                    self.__SheetBoot.cell(row= r, column= int(VarPositionTitle[self.__Promissed])).value.day
                    datePD = str(self.__SheetBoot.cell(row= r, column= int(VarPositionTitle[self.__Promissed])).value)
                except:
                    datePD = ""
                try:
                    self.__SheetBoot.cell(row= r, column= int(VarPositionTitle[self.__Assigned])).value.day
                    dateAD = str(self.__SheetBoot.cell(row= r, column= int(VarPositionTitle[self.__Assigned])).value)
                except:
                    dateAD = ""
                if(self.__Promissed!='Promised Date'):
                    Data['Info'].append({
                        'po_nomber': Po,
                        'Cut': Cut,
                        self.__Promissed: datePD,
                        'assigned_date': dateAD,
                        'NumAssigned': NumAssigned
                        })
                else:
                    Data['Info'].append({
                        'po_nomber': Po,
                        self.__Promissed: datePD
                        })
        return Data

    async def ModelData(self):
        data = self.AddInfo()
        df = pd.json_normalize(data['Info'])
        if(self.__Promissed!='Promised Date'):
            df = df.groupby(['po_nomber','Cut']).max()
        else:
            df = df.groupby(['po_nomber']).max()
        df = df.reset_index()
        return df
